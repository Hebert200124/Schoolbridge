from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, g, abort
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Student, Subject, StudentSubject, MonthlyTest, ExamMark, FeeAccount, Payment, FeeSetting, Timetable, ExamTimetable, PrincipalComment, TeacherRemark, ActivityLog, Activity, StaffLeave, OTPCode, Campus, zim_grade
from config import Config
from functools import wraps
from datetime import datetime, date, timedelta
import os
import re
import random
import string
import traceback
import sib_api_v3_sdk
from sib_api_v3_sdk.rest import ApiException
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy.exc import IntegrityError
from email_validator import validate_email, EmailNotValidError

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    _Limiter = Limiter
    _get_remote_address = get_remote_address
except ImportError:
    _Limiter = None
    _get_remote_address = None


class _NoopLimiter:
    def limit(self, *args, **kwargs):
        def decorator(fn):
            return fn
        return decorator

app = Flask(__name__)
app.config.from_object(Config)
app.jinja_env.auto_reload = True
app.config['TEMPLATES_AUTO_RELOAD'] = True

app.config['RATELIMIT_ENABLED'] = os.environ.get('RATELIMIT_ENABLED', 'true').lower() == 'true'
if _Limiter is not None:
    limiter = _Limiter(key_func=_get_remote_address, app=app)
else:
    limiter = _NoopLimiter()


CAMPUS_SCOPED_TABLES = [
    'users', 'students', 'subjects', 'monthly_tests', 'exam_marks',
    'principal_comments', 'fee_accounts', 'payments', 'fee_settings',
    'teacher_remarks', 'timetables', 'exam_timetables', 'activity_logs',
    'activities', 'staff_leaves',
]


def _ensure_campus_columns():
    """Idempotently add campus_id to existing tables and backfill to Main Campus.

    Runs on every boot so existing databases (local SQLite and production
    Postgres on Render) gain the multi-campus columns without downtime.
    Fresh databases already get them from db.create_all(). The standalone
    migrate_campuses.py script additionally adds the FK constraints and swaps
    the old global unique constraints for per-campus ones on Postgres.
    """
    try:
        if 'campuses' not in inspect(db.engine).get_table_names():
            return
        main = Campus.query.filter_by(code='MAIN').first()
        if not main:
            main = Campus(name='Main Campus', code='MAIN', address='')
            db.session.add(main)
            db.session.flush()
        for table in CAMPUS_SCOPED_TABLES:
            inspector = inspect(db.engine)
            if table not in inspector.get_table_names():
                continue
            cols = {c['name'] for c in inspector.get_columns(table)}
            if 'campus_id' in cols:
                continue
            db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN campus_id INTEGER'))
            db.session.execute(text(f'UPDATE {table} SET campus_id = :cid WHERE campus_id IS NULL'),
                               {'cid': main.id})
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        print(f'[migration] campus columns skipped: {exc}')


def _ensure_per_campus_uniques():
    """Swap old global unique constraints for per-campus ones (Postgres only).

    Databases created before multi-campus support have a GLOBAL UNIQUE on
    subjects.code (and fee_settings.form). Seeding standard subjects for a
    second campus then violates it and crashes. This drops the old constraint
    and adds the per-campus (campus_id, code) / (campus_id, form) one, so the
    fix applies automatically at boot without running migrate_campuses.py.
    """
    try:
        if db.engine.dialect.name != 'postgresql':
            return
        if 'campuses' not in inspect(db.engine).get_table_names():
            return
        for table, column, new_name in [
            ('subjects', 'code', 'uq_subjects_campus_code'),
            ('fee_settings', 'form', 'uq_fee_settings_campus_form'),
        ]:
            try:
                _migrate_table_uniques(table, column, new_name)
            except Exception as exc:
                print(f'[migration] {table} per-campus unique skipped: {exc}')
    except Exception as exc:
        print(f'[migration] per-campus uniques skipped: {exc}')


def _migrate_table_uniques(table, column, new_name):
    """Drop the legacy global unique on one column and add a per-campus one.

    Each table runs in its own transaction so a failure on one never rolls
    back the other. Postgres stores a column-level `unique=True` as the unique
    constraint/index `subjects_code_key`; it can be a constraint, a standalone
    index, or both, so this inspects the pg catalogs directly instead of
    relying on SQLAlchemy's constraint reflection.
    """
    with db.engine.begin() as conn:
        inspector = inspect(db.engine)
        if table not in inspector.get_table_names():
            return
        cols = {c['name'] for c in inspector.get_columns(table)}
        if 'campus_id' not in cols:
            return

        # 1) Drop a unique CONSTRAINT whose only column is `column`.
        #    DROP CONSTRAINT also removes its backing index.
        const_rows = conn.execute(text('''
            SELECT c.conname
            FROM pg_constraint c
            JOIN pg_class t ON t.oid = c.conrelid
            JOIN pg_namespace n ON n.oid = t.relnamespace
            WHERE n.nspname = current_schema()
              AND t.relname = :t
              AND c.contype = 'u'
              AND array_length(c.conkey, 1) = 1
              AND c.conkey::int[] = (
                  SELECT ARRAY[attnum] FROM pg_attribute a
                  WHERE a.attrelid = t.oid AND a.attname = :col)
        '''), {'t': table, 'col': column}).fetchall()
        for (conname,) in const_rows:
            print(f'[migration] drop global unique constraint {conname} on {table}')
            conn.execute(text(f'ALTER TABLE {table} DROP CONSTRAINT IF EXISTS "{conname}"'))

        # 2) Drop a standalone unique INDEX on `column` (no backing constraint).
        #    Runs after the constraint drop so constraint-backed indexes are
        #    already gone and cannot trigger the "constraint requires it" error.
        idx_rows = conn.execute(text('''
            SELECT i.relname
            FROM pg_index x
            JOIN pg_class i ON i.oid = x.indexrelid
            JOIN pg_class t ON t.oid = x.indrelid
            JOIN pg_namespace n ON n.oid = t.relnamespace
            WHERE n.nspname = current_schema()
              AND t.relname = :t
              AND x.indisunique
              AND x.indnkeyatts = 1
              AND (SELECT a.attname FROM pg_attribute a
                   WHERE a.attrelid = t.oid AND a.attnum = x.indkey[0]) = :col
        '''), {'t': table, 'col': column}).fetchall()
        for (idxname,) in idx_rows:
            print(f'[migration] drop global unique index {idxname} on {table}')
            conn.execute(text(f'DROP INDEX IF EXISTS "{idxname}"'))

        # 3) Add the per-campus unique constraint if it is not already there.
        existing = {uc['name'] for uc in inspector.get_unique_constraints(table)}
        if new_name not in existing:
            print(f'[migration] add per-campus unique {new_name} on {table}')
            conn.execute(text(
                f'ALTER TABLE {table} ADD CONSTRAINT {new_name} '
                f'UNIQUE (campus_id, {column}) NOT VALID'))
            conn.execute(text(f'ALTER TABLE {table} VALIDATE CONSTRAINT {new_name}'))

db.init_app(app)
with app.app_context():
    db.create_all()
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    existing_columns = [col['name'] for col in inspector.get_columns('students')]
    if 'curriculum' not in existing_columns:
        with db.engine.connect() as conn:
            conn.execute(text('ALTER TABLE students ADD COLUMN curriculum VARCHAR(20)'))
            conn.commit()

    # otp_codes.phone stores the email address now; widen it on Postgres (VARCHAR(20) is too short)
    if db.engine.dialect.name == 'postgresql':
        phone_cols = [c for c in inspector.get_columns('otp_codes') if c['name'] == 'phone']
        if phone_cols:
            length = getattr(phone_cols[0]['type'], 'length', None)
            if length is not None and length < 255:
                with db.engine.connect() as conn:
                    conn.execute(text('ALTER TABLE otp_codes ALTER COLUMN phone TYPE VARCHAR(255)'))
                    conn.commit()

    _ensure_campus_columns()
    _ensure_per_campus_uniques()



login_manager = LoginManager(app)
login_manager.login_view = 'auth_login'
login_manager.login_message_category = 'info'


@app.context_processor
def inject_now():
    return {'now': datetime.now}


def generate_student_id():
    year = datetime.now().year
    prefix = str(year)
    max_num = 0
    for s in Student.query.all():
        sid = s.student_id
        if sid and sid.startswith(prefix) and sid[len(prefix):].isdigit():
            max_num = max(max_num, int(sid[len(prefix):]))
    return f'{prefix}{max_num + 1:03d}'


def log_activity(action, description=None, user=None, student=None, visibility='public', campus_id=None):
    if campus_id is None:
        campus_id = (getattr(user, 'campus_id', None) or getattr(student, 'campus_id', None)
                     or getattr(g, 'current_campus_id', None))
    entry = ActivityLog(action=action, description=description, visibility=visibility,
                        campus_id=campus_id,
                        user_id=user.id if user else None,
                        student_id=student.id if student else None)
    db.session.add(entry)
    db.session.commit()


def staff_email(username, email):
    """Users require a non-null, unique email; generate a placeholder when blank."""
    email = (email or '').strip()
    return email if email else f'{username.lower()}@schoolbridge.local'


def generate_receipt():
    while True:
        ref = 'RCP' + ''.join(random.choices(string.digits, k=8))
        if not Payment.query.filter_by(receipt_number=ref).first():
            return ref


def get_current_term():
    return 'Term 1'


def get_term_fee(form, campus_id=None):
    campus_id = campus_id or getattr(g, 'current_campus_id', None)
    fs = FeeSetting.query.filter_by(form=form, campus_id=campus_id).first()
    return fs.term_fee if fs else 0.0


def get_detected_fee(student, term=None):
    term = term or get_current_term()
    fee_account = FeeAccount.query.filter_by(student_id=student.id, term=term).first()
    if fee_account:
        return fee_account.total_fees
    return get_term_fee(student.form, campus_id=student.campus_id)


def ensure_fee_account(student, term=None, amount_paid=0.0):
    term = term or get_current_term()
    fee_account = FeeAccount.query.filter_by(student_id=student.id, term=term).first()
    if not fee_account:
        total = get_term_fee(student.form, campus_id=student.campus_id)
        fee_account = FeeAccount(student_id=student.id, campus_id=student.campus_id, term=term,
                                 total_fees=total, amount_paid=amount_paid, balance=total - amount_paid)
        db.session.add(fee_account)
    return fee_account


def compute_pass_rates(term, campus_id=None):
    """Compute per-subject and per-form pass rates with 3 queries total.

    Previously this looped over every subject and every form, firing one query
    per iteration (the classic N+1 pattern on the dashboards). Scoped to a
    single campus; None means all campuses (super_admin).
    """
    if campus_id is None:
        campus_id = getattr(g, 'current_campus_id', None)
    subjects = Subject.query.filter_by(campus_id=campus_id).all() if campus_id else Subject.query.all()
    exams = (ExamMark.query.filter_by(term=term, campus_id=campus_id).all() if campus_id
             else ExamMark.query.filter_by(term=term).all())

    by_subject = {}
    for e in exams:
        by_subject.setdefault(e.subject_id, []).append(e)

    pass_rates = []
    for subj in subjects:
        group = by_subject.get(subj.id)
        if group:
            passed = sum(1 for e in group if e.total_marks > 0 and e.marks / e.total_marks * 100 >= 50)
            pass_rates.append({'subject': subj.name, 'rate': round(passed / len(group) * 100, 1), 'total': len(group)})

    students = Student.query.filter_by(campus_id=campus_id).all() if campus_id else Student.query.all()
    student_form = {s.id: s.form for s in students}
    by_form = {}
    for e in exams:
        f = student_form.get(e.student_id)
        if f:
            by_form.setdefault(f, []).append(e)

    form_pass_rates = []
    for f in ['Form 1', 'Form 2', 'Form 3', 'Form 4', 'Form 5', 'Form 6']:
        group = by_form.get(f)
        if group:
            passed = sum(1 for e in group if e.total_marks > 0 and e.marks / e.total_marks * 100 >= 50)
            form_pass_rates.append({'form': f, 'rate': round(passed / len(group) * 100, 1), 'total': len(group)})

    return pass_rates, form_pass_rates


def compute_expected_fees(term, campus_id):
    """Expected fees = one term fee per student (their account total if set,
    otherwise the campus default for their form).

    Counts every student on the campus (active or transferred) so the
    expected figure matches the total the school is entitled to collect.
    Previously this was the raw sum of FeeAccount.total_fees, which could
    overcount when a student had duplicate accounts for the same term.
    """
    accounts = FeeAccount.query.filter(FeeAccount.term == term, FeeAccount.campus_id == campus_id).all()
    by_student = {}
    for a in accounts:
        by_student[a.student_id] = a.total_fees
    total = 0.0
    for s in Student.query.filter_by(campus_id=campus_id).all():
        total += by_student.get(s.id, get_term_fee(s.form, campus_id=campus_id))
    return total


def staff_currently_on_leave(campus_id, today=None):
    """Distinct staff currently on approved leave that overlaps `today`."""
    today = today or date.today()
    rows = (StaffLeave.query.filter(StaffLeave.campus_id == campus_id,
                                    StaffLeave.status == 'Approved',
                                    StaffLeave.start_date <= today,
                                    StaffLeave.end_date >= today)
            .with_entities(StaffLeave.user_id).distinct().all())
    return {r[0] for r in rows}


@login_manager.user_loader
def load_user(user_id):
    if user_id.startswith('student_'):
        return Student.query.get(int(user_id.replace('student_', '')))
    return User.query.get(int(user_id))


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('auth_login'))
            if isinstance(current_user, Student):
                flash('Access denied. Staff only.', 'danger')
                return redirect(url_for('student_dashboard'))
            if current_user.role not in roles and current_user.role != 'super_admin':
                flash('Access denied.', 'danger')
                return redirect(url_for('staff_dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth_login'))
        if not isinstance(current_user, User) or getattr(current_user, 'role', '') != 'super_admin':
            flash('Access denied. Super admin only.', 'danger')
            return redirect(url_for('staff_dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth_login'))
        if not isinstance(current_user, Student):
            flash('Access denied. Students only.', 'danger')
            return redirect(url_for('staff_dashboard'))
        return f(*args, **kwargs)
    return decorated_function


@app.before_request
def _set_campus_context():
    """Expose the current user's campus to the scoped() query helper."""
    if current_user.is_authenticated:
        g.current_campus_id = getattr(current_user, 'campus_id', None)
        g.is_super_admin = getattr(current_user, 'role', '') == 'super_admin'
    else:
        g.current_campus_id = None
        g.is_super_admin = False


def scoped(model):
    """Return a query over `model` limited to the logged-in user's campus.

    - super_admin sees every campus (no filter applied).
    - Staff/students only ever see rows where campus_id matches theirs.
    - Outside a request (login lookups, seeding, migrations) no filter is
      applied so global operations keep working.
    """
    if getattr(g, 'is_super_admin', False):
        return model.query
    cid = getattr(g, 'current_campus_id', None)
    if cid is not None and hasattr(model, 'campus_id'):
        return model.query.filter(model.campus_id == cid)
    return model.query


def scoped_get_or_404(model, obj_id):
    """get_or_404 that respects campus isolation (other-campus ids 404)."""
    obj = scoped(model).filter(model.id == obj_id).first()
    if obj is None:
        abort(404)
    return obj


def get_selected_campus_id():
    """Campus that new records should be attached to.

    super_admin may pick any campus via the `campus_id` form field; everyone
    else is always bound to their own campus.
    """
    if getattr(g, 'is_super_admin', False):
        cid = request.form.get('campus_id', type=int)
        if cid and Campus.query.get(cid):
            return cid
    return getattr(current_user, 'campus_id', None)


# ============ AUTH ROUTES ============

@app.route('/')
def index():
    if current_user.is_authenticated:
        if isinstance(current_user, Student):
            return redirect(url_for('student_dashboard'))
        if getattr(current_user, 'role', '') == 'super_admin':
            return redirect(url_for('super_admin_dashboard'))
        return redirect(url_for('staff_dashboard'))
    return redirect(url_for('auth_login'))


@app.route('/auth/login', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def auth_login():
    if current_user.is_authenticated:
        if isinstance(current_user, Student):
            return redirect(url_for('student_dashboard'))
        if getattr(current_user, 'role', '') == 'super_admin':
            return redirect(url_for('super_admin_dashboard'))
        return redirect(url_for('staff_dashboard'))

    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''

        student = Student.query.filter_by(student_id=username).first()
        if student and student.check_password(password):
            if not student.is_active:
                flash('Account deactivated. Contact admin.', 'danger')
                return render_template('auth/login.html')
            login_user(student)
            flash(f'Welcome {student.full_name}!', 'success')
            return redirect(url_for('student_dashboard'))

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            if not user.is_active:
                flash('Account deactivated. Contact principal.', 'danger')
                return render_template('auth/login.html')
            login_user(user)
            flash(f'Welcome {user.full_name}!', 'success')
            if user.role == 'super_admin':
                return redirect(url_for('super_admin_dashboard'))
            return redirect(url_for('staff_dashboard'))

        flash('Invalid username or password', 'danger')

    return render_template('auth/login.html')


@app.route('/auth/logout')
@login_required
def auth_logout():
    logout_user()
    flash('You have been logged out', 'info')
    return redirect(url_for('auth_login'))


@app.route('/auth/change-password', methods=['GET', 'POST'])
@login_required
def auth_change_password():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        if not current_user.check_password(current_pw):
            flash('Current password is incorrect.', 'danger')
            return render_template('auth/change_password.html')

        errors = password_strength_errors(new_pw)
        if errors:
            flash(errors[0], 'danger')
            return render_template('auth/change_password.html')

        if current_pw == new_pw:
            flash('New password must be different from the current password.', 'danger')
            return render_template('auth/change_password.html')

        if new_pw != confirm_pw:
            flash('New passwords do not match.', 'danger')
            return render_template('auth/change_password.html')

        current_user.set_password(new_pw)
        db.session.commit()
        flash('Password changed successfully! Please log in with your new password.', 'success')
        return redirect(url_for('auth_logout'))

    return render_template('auth/change_password.html')


def send_otp_email(to_email, otp_code, expires_minutes=10):
    """Send the OTP code to the user's email via Brevo transactional email API."""
    try:
        configuration = sib_api_v3_sdk.Configuration()
        configuration.api_key['api-key'] = app.config['BREVO_API_KEY']
        api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(configuration))

        html_content = f"""
        <div style="font-family:Arial,Helvetica,sans-serif;max-width:480px;margin:0 auto;border:1px solid #E2E8F0;border-radius:10px;overflow:hidden;">
            <div style="background:#334E68;color:#ffffff;padding:24px;text-align:center;">
                <h2 style="margin:0;">SchoolBridge</h2>
                <p style="margin:4px 0 0;opacity:0.85;">Password Reset</p>
            </div>
            <div style="padding:24px;">
                <p>Hello,</p>
                <p>Use the one-time password below to reset your SchoolBridge account password. This code expires in <strong>{expires_minutes} minutes</strong>.</p>
                <div style="text-align:center;font-size:32px;font-weight:700;letter-spacing:8px;color:#334E68;background:#F9FAFB;border-radius:8px;padding:16px;margin:16px 0;">{otp_code}</div>
                <p style="color:#6B7280;font-size:13px;">If you did not request this, you can safely ignore this email.</p>
            </div>
        </div>
        """

        send_smtp_email = sib_api_v3_sdk.SendSmtpEmail(
            to=[{'email': to_email}],
            sender={'name': 'SchoolBridge', 'email': app.config['MAIL_DEFAULT_SENDER']},
            subject='SchoolBridge: Your Password Reset OTP',
            html_content=html_content
        )
        api_instance.send_transac_email(send_smtp_email)
        app.logger.info(f'[BREVO] OTP sent to: {to_email}')
        return True
    except ApiException as e:
        app.logger.error(f'[BREVO] Failed to send OTP email to {to_email}: {e}')
        app.logger.error(f'[BREVO] Response body: {e.body}')
        return False
    except Exception as e:
        app.logger.error(f'[BREVO] Unexpected error sending OTP email to {to_email}: {e}')
        return False


def _reg_variants(reg_raw):
    reg_clean = reg_raw.replace('-', '').replace('REG', '').replace('reg', '')
    return list(set([
        reg_raw,
        reg_raw.replace('-', ''),
        f'REG-{reg_clean}',
        f'REG{reg_clean}',
        reg_raw.upper(),
    ]))


def find_person_by_reg(reg_raw):
    variants = _reg_variants(reg_raw)
    candidates = list(User.query.filter(User.reg_number.in_(variants)).all())
    candidates += list(Student.query.filter(Student.reg_number.in_(variants)).all())
    candidates += list(Student.query.filter(Student.student_id.in_(variants)).all())
    candidates += list(User.query.filter(User.username.in_(variants)).all())
    seen = set()
    for p in candidates:
        if p.id in seen:
            continue
        seen.add(p.id)
        return p
    return None


@app.route('/auth/reset-password', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def auth_reset_password():
    if request.method == 'POST':
        reg_number = request.form.get('reg_number', '').strip()
        if not reg_number:
            flash('Please enter your registration number.', 'warning')
            return redirect(url_for('auth_reset_password'))
        person = find_person_by_reg(reg_number)

        session['reset_reg_number'] = reg_number
        if person:
            session['reset_user_type'] = 'User' if isinstance(person, User) else 'Student'
            session['reset_user_id'] = person.id
        else:
            session.pop('reset_user_type', None)
            session.pop('reset_user_id', None)

        return redirect(url_for('auth_reset_email'))

    return render_template('auth/reset_password.html')


@app.route('/auth/reset-email', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def auth_reset_email():
    reg_number = session.get('reset_reg_number')
    if not reg_number:
        flash('Please enter your registration number first.', 'warning')
        return redirect(url_for('auth_reset_password'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        user_type = session.get('reset_user_type')
        user_id = session.get('reset_user_id')

        if not is_valid_email(email):
            flash('If those details match our records, an OTP has been sent to your email.', 'info')
            return redirect(url_for('auth_reset_email'))

        person = None
        if user_type and user_id:
            person = User.query.get(user_id) if user_type == 'User' else Student.query.get(user_id)

        if not person or (person.email or '').strip().lower() != email.strip().lower():
            flash('If those details match our records, an OTP has been sent to your email.', 'info')
            return redirect(url_for('auth_reset_email'))

        email_clean = (person.email or '').strip().lower()

        latest = OTPCode.query.filter_by(
            user_type=user_type, user_id=user_id, used=False
        ).order_by(OTPCode.created_at.desc()).first()

        if latest and latest.is_rate_limited():
            remaining = int(900 - (datetime.utcnow() - latest.request_window_start).total_seconds())
            flash(f'Too many OTP requests. Try again in {remaining // 60} minute(s).', 'danger')
            return redirect(url_for('auth_reset_email'))

        if latest and (datetime.utcnow() - latest.created_at).total_seconds() < 60:
            flash('Please wait at least 1 minute before requesting a new OTP.', 'warning')
            return redirect(url_for('auth_reset_email'))

        code = f'{random.randint(0, 999999):06d}'
        now = datetime.utcnow()

        if latest:
            latest.request_count += 1
            otp = latest
        else:
            otp = OTPCode(
                user_type=user_type, user_id=user_id,
                phone=email_clean, code=code,
                created_at=now, expires_at=now + timedelta(minutes=10),
                request_window_start=now
            )
            db.session.add(otp)

        otp.code = code
        otp.created_at = now
        otp.expires_at = now + timedelta(minutes=10)
        otp.used = False
        otp.attempts = 0

        db.session.commit()

        send_otp_email(email_clean, code)

        if app.debug:
            flash(f'[DEV] OTP for {email_clean}: {code}', 'info')

        session['reset_otp_id'] = otp.id
        session['reset_phone'] = email_clean
        session['reset_user_type'] = user_type
        session['reset_user_id'] = user_id

        flash('A 6-digit OTP has been sent to your email.', 'success')
        return redirect(url_for('auth_reset_code'))

    return render_template('auth/reset_email.html', reg_number=reg_number)


@app.route('/auth/reset-code', methods=['GET', 'POST'])
@limiter.limit('10 per minute', methods=['POST'])
def auth_reset_code():
    otp_id = session.get('reset_otp_id')
    if not otp_id:
        flash('Please request a password reset first.', 'warning')
        return redirect(url_for('auth_reset_password'))

    otp = OTPCode.query.get(otp_id)
    if not otp or otp.used:
        flash('OTP has already been used. Please request a new one.', 'warning')
        return redirect(url_for('auth_reset_password'))

    if otp.is_expired():
        flash('OTP has expired. Please request a new one.', 'danger')
        return redirect(url_for('auth_reset_password'))

    if request.method == 'POST':
        code = (request.form.get('code', '') or '').strip()

        if not re.fullmatch(r'\d{6}', code):
            flash('Invalid OTP format.', 'danger')
            return render_template('auth/reset_code.html')

        if otp.attempts >= 5:
            flash('Too many failed attempts. Please request a new OTP.', 'danger')
            return redirect(url_for('auth_reset_password'))

        if code != otp.code:
            otp.attempts += 1
            db.session.commit()
            remaining = 5 - otp.attempts
            flash(f'Invalid OTP. {remaining} attempt(s) remaining.', 'danger')
            return render_template('auth/reset_code.html')

        otp.used = True
        db.session.commit()
        session['reset_code_verified'] = True

        flash('OTP verified. Please set your new password.', 'success')
        return redirect(url_for('auth_reset_set_password'))

    return render_template('auth/reset_code.html')


@app.route('/auth/reset-set-password', methods=['GET', 'POST'])
def auth_reset_set_password():
    if not session.get('reset_code_verified'):
        flash('Please verify your OTP first.', 'warning')
        return redirect(url_for('auth_reset_code'))

    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')

        errors = password_strength_errors(new_password)
        if errors:
            flash(errors[0], 'danger')
            return render_template('auth/reset_set_password.html')

        if new_password != confirm:
            flash('Passwords do not match.', 'danger')
            return render_template('auth/reset_set_password.html')

        user_type = session.get('reset_user_type')
        user_id = session.get('reset_user_id')
        if user_type == 'User':
            user = User.query.get(user_id)
            if user:
                user.set_password(new_password)
        elif user_type == 'Student':
            student = Student.query.get(user_id)
            if student:
                student.set_password(new_password)
        db.session.commit()

        for k in ('reset_otp_id', 'reset_phone', 'reset_user_type', 'reset_user_id', 'reset_code_verified'):
            session.pop(k, None)

        flash('Password reset successfully! You can now log in.', 'success')
        return redirect(url_for('auth_login'))

    return render_template('auth/reset_set_password.html')


# ============ STAFF DASHBOARD ============

@app.route('/staff/dashboard')
@login_required
def staff_dashboard():
    if isinstance(current_user, Student):
        return redirect(url_for('student_dashboard'))
    if getattr(current_user, 'role', '') == 'super_admin':
        return redirect(url_for('super_admin_dashboard'))

    stats = {}
    if current_user.role == 'teacher':
        subj = current_user.teacher_subject
        stats['subject_name'] = subj.name if subj else 'Not assigned'
        stats['student_count'] = StudentSubject.query.filter_by(subject_id=current_user.subject_id).count() if current_user.subject_id else 0
        stats['activities_count'] = scoped(Activity).count()
        stats['recent_activity_posts'] = scoped(Activity).order_by(Activity.created_at.desc()).limit(10).all()
    elif current_user.role == 'cashier':
        stats['pending_clearance'] = scoped(Payment).filter_by(cleared=False).count()
        stats['total_payments_today'] = scoped(Payment).filter(db.func.date(Payment.created_at) == date.today()).count()
        stats['recent_activity_posts'] = scoped(Activity).order_by(Activity.created_at.desc()).limit(10).all()
    elif current_user.role == 'admin':
        stats['total_students'] = scoped(Student).count()
        stats['active_students'] = scoped(Student).filter_by(is_active=True).count()
        stats['staff_count'] = scoped(User).filter(User.role.notin_(['principal', 'super_admin'])).count()
        term = 'Term 1'
        stats['term'] = term
        cid = current_user.campus_id
        stats['total_expected'] = compute_expected_fees(term, cid)
        stats['total_collected'] = db.session.query(db.func.sum(Payment.amount)).filter(Payment.cleared == True, Payment.campus_id == cid).scalar() or 0
        fee_accounts = FeeAccount.query.filter_by(term=term, campus_id=cid).all()
        stats['outstanding'] = sum(fa.balance for fa in fee_accounts if fa.balance > 0)
        stats['outstanding_list'] = FeeAccount.query.filter(FeeAccount.term == term, FeeAccount.balance > 0, FeeAccount.campus_id == cid).order_by(FeeAccount.balance.desc()).options(joinedload(FeeAccount.student)).all()
        now_date = date.today()
        stats['upcoming_exams'] = scoped(ExamTimetable).filter(ExamTimetable.exam_date >= now_date).order_by(ExamTimetable.exam_date).options(joinedload(ExamTimetable.subject_rel)).limit(5).all()
        stats['recent_activities'] = scoped(ActivityLog).options(joinedload(ActivityLog.user), joinedload(ActivityLog.student)).order_by(ActivityLog.created_at.desc()).limit(10).all()
        stats['recent_activity_posts'] = scoped(Activity).options(joinedload(Activity.creator)).order_by(Activity.created_at.desc()).limit(10).all()
        stats['staff_on_leave'] = len(staff_currently_on_leave(cid, now_date))
        stats['pass_rates'], stats['form_pass_rates'] = compute_pass_rates(term, campus_id=cid)
        staff_leaves = stats['staff_on_leave']
        stats['active_staff_count'] = stats['staff_count'] - staff_leaves
        stats['staff_on_leave_count'] = staff_leaves
    elif current_user.role == 'principal':
        term = 'Term 1'
        stats['term'] = term
        cid = current_user.campus_id
        stats['total_staff'] = scoped(User).filter(User.role.notin_(['principal', 'super_admin'])).count()
        stats['total_students'] = scoped(Student).count()
        stats['active_students'] = scoped(Student).filter_by(is_active=True).count()
        stats['total_expected'] = compute_expected_fees(term, cid)
        stats['total_collected'] = db.session.query(db.func.sum(Payment.amount)).filter(Payment.cleared == True, Payment.campus_id == cid).scalar() or 0
        fee_accounts = FeeAccount.query.filter_by(term=term, campus_id=cid).all()
        stats['outstanding'] = sum(fa.balance for fa in fee_accounts if fa.balance > 0)
        stats['outstanding_list'] = FeeAccount.query.filter(FeeAccount.term == term, FeeAccount.balance > 0, FeeAccount.campus_id == cid).order_by(FeeAccount.balance.desc()).options(joinedload(FeeAccount.student)).all()
        active_staff = scoped(User).filter(User.role.notin_(['principal', 'super_admin']), User.is_active == True).count()
        now_date = date.today()
        staff_on_leave = len(staff_currently_on_leave(cid, now_date))
        stats['staff_on_leave'] = staff_on_leave
        stats['active_staff_count'] = active_staff
        stats['staff_on_leave_count'] = staff_on_leave
        stats['upcoming_exams'] = scoped(ExamTimetable).filter(ExamTimetable.exam_date >= now_date).order_by(ExamTimetable.exam_date).options(joinedload(ExamTimetable.subject_rel)).limit(5).all()
        stats['recent_activities'] = scoped(ActivityLog).options(joinedload(ActivityLog.user), joinedload(ActivityLog.student)).order_by(ActivityLog.created_at.desc()).limit(10).all()
        stats['recent_activity_posts'] = scoped(Activity).options(joinedload(Activity.creator)).order_by(Activity.created_at.desc()).limit(10).all()
        stats['pass_rates'], stats['form_pass_rates'] = compute_pass_rates(term, campus_id=cid)

    return render_template('staff/dashboard.html', stats=stats)


@app.route('/staff/activities')
@login_required
def staff_activities():
    activities = scoped(ActivityLog).order_by(ActivityLog.created_at.desc()).all()
    return render_template('staff/activities.html', activities=activities)


# ============ TEACHER ROUTES ============

@app.route('/staff/teacher/class')
@login_required
@role_required('teacher')
def teacher_class():
    if not current_user.subject_id:
        flash('No subject assigned.', 'warning')
        return redirect(url_for('staff_dashboard'))

    subject = scoped_get_or_404(Subject, current_user.subject_id)
    student_subjects = StudentSubject.query.filter_by(subject_id=current_user.subject_id).all()
    students = [ss.student for ss in student_subjects if ss.student.is_active and ss.student.campus_id == current_user.campus_id]
    recent_activities = scoped(ActivityLog).order_by(ActivityLog.created_at.desc()).limit(5).all()

    return render_template('staff/teacher/class.html', subject=subject, students=students, recent_activities=recent_activities)


@app.route('/staff/teacher/student/<int:student_id>/results')
@login_required
@role_required('teacher')
def teacher_student_results(student_id):
    student = scoped_get_or_404(Student, student_id)
    monthly_tests = MonthlyTest.query.filter_by(student_id=student.id, subject_id=current_user.subject_id).order_by(MonthlyTest.academic_year.desc(), MonthlyTest.term).all()
    exam_marks = ExamMark.query.filter_by(student_id=student.id, subject_id=current_user.subject_id).order_by(ExamMark.academic_year.desc(), ExamMark.term).all()
    subject = scoped_get_or_404(Subject, current_user.subject_id)
    return render_template('staff/teacher/student_results.html', student=student, monthly_tests=monthly_tests, exam_marks=exam_marks, subject=subject)


@app.route('/staff/teacher/marks')
@login_required
@role_required('teacher')
def teacher_marks():
    if not current_user.subject_id:
        flash('No subject assigned.', 'warning')
        return redirect(url_for('staff_dashboard'))

    subject = scoped_get_or_404(Subject, current_user.subject_id)
    student_subjects = StudentSubject.query.filter_by(subject_id=current_user.subject_id).all()
    students = [ss.student for ss in student_subjects if ss.student.is_active and ss.student.campus_id == current_user.campus_id]
    monthly_tests = MonthlyTest.query.filter_by(subject_id=current_user.subject_id).all()
    exam_marks_list = ExamMark.query.filter_by(subject_id=current_user.subject_id).all()

    marks_data = {}
    for s in students:
        s_monthly = [mt for mt in monthly_tests if mt.student_id == s.id]
        s_exam = [em for em in exam_marks_list if em.student_id == s.id]
        marks_data[s.id] = {'monthly': s_monthly, 'exam': s_exam}

    return render_template('staff/teacher/marks.html', subject=subject, students=students, marks_data=marks_data)


@app.route('/staff/teacher/marks/add', methods=['POST'])
@login_required
@role_required('teacher')
def teacher_add_mark():
    student_id = request.form.get('student_id')
    mark_type = request.form.get('mark_type')
    term = request.form.get('term')
    month = request.form.get('month')
    exam_type = request.form.get('exam_type')
    marks = float(request.form.get('marks'))
    total_marks = float(request.form.get('total_marks', 100))
    academic_year = request.form.get('academic_year', str(datetime.now().year))
    comment = request.form.get('comment')

    student = scoped(Student).filter_by(id=student_id).first()
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('teacher_marks'))

    if mark_type == 'monthly':
        test = MonthlyTest(
            student_id=student.id, subject_id=current_user.subject_id,
            campus_id=current_user.campus_id,
            term=term, academic_year=academic_year, month=month,
            marks=marks, total_marks=total_marks, teacher_id=current_user.id,
            comment=comment or None
        )
        db.session.add(test)
    elif mark_type == 'exam':
        exam = ExamMark(
            student_id=student.id, subject_id=current_user.subject_id,
            campus_id=current_user.campus_id,
            term=term, academic_year=academic_year, exam_type=exam_type,
            marks=marks, total_marks=total_marks, teacher_id=current_user.id,
            comment=comment or None
        )
        db.session.add(exam)

    db.session.commit()
    log_activity('Marks entered', f'{mark_type} - {student.full_name}: {marks}/{total_marks}', user=current_user, student=student)
    flash(f'Marks added! Grade: {zim_grade(marks, total_marks, "A" if "6" in student.form or "5" in student.form else "O")}', 'success')
    return redirect(url_for('teacher_marks'))


@app.route('/staff/teacher/marks/edit/<int:mark_id>', methods=['POST'])
@login_required
@role_required('teacher')
def teacher_edit_mark(mark_id):
    mark_type = request.form.get('mark_type')
    marks = float(request.form.get('marks'))
    total_marks = float(request.form.get('total_marks', 100))
    comment = request.form.get('comment')

    if mark_type == 'monthly':
        mark = scoped_get_or_404(MonthlyTest, mark_id)
        if mark.subject_id != current_user.subject_id:
            flash('Access denied.', 'danger')
            return redirect(url_for('teacher_marks'))
        mark.marks = marks
        mark.total_marks = total_marks
        mark.comment = comment or None
    elif mark_type == 'exam':
        mark = scoped_get_or_404(ExamMark, mark_id)
        if mark.subject_id != current_user.subject_id:
            flash('Access denied.', 'danger')
            return redirect(url_for('teacher_marks'))
        mark.marks = marks
        mark.total_marks = total_marks
        mark.comment = comment or None

    db.session.commit()
    flash('Marks updated!', 'success')
    return redirect(url_for('teacher_marks'))


@app.route('/staff/teacher/marks/delete/<int:mark_id>/<mark_type>')
@login_required
@role_required('teacher')
def teacher_delete_mark(mark_id, mark_type):
    if mark_type == 'monthly':
        mark = scoped_get_or_404(MonthlyTest, mark_id)
        if mark.teacher_id != current_user.id:
            flash('Access denied.', 'danger')
            return redirect(url_for('teacher_marks'))
        db.session.delete(mark)
    elif mark_type == 'exam':
        mark = scoped_get_or_404(ExamMark, mark_id)
        if mark.teacher_id != current_user.id:
            flash('Access denied.', 'danger')
            return redirect(url_for('teacher_marks'))
        db.session.delete(mark)

    db.session.commit()
    flash('Mark deleted.', 'success')
    return redirect(url_for('teacher_marks'))


@app.route('/staff/teacher/remark/save', methods=['POST'])
@login_required
@role_required('teacher')
def teacher_save_remark():
    student_id = request.form.get('student_id')
    remark = request.form.get('remark')
    student = scoped(Student).filter_by(id=student_id).first()
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('teacher_class'))

    tr = TeacherRemark(student_id=student.id, teacher_id=current_user.id, remark=remark, campus_id=current_user.campus_id)
    db.session.add(tr)
    db.session.commit()
    flash(f'Remark added for {student.full_name}.', 'success')
    return redirect(url_for('teacher_class'))


# ============ CASHIER ROUTES ============

@app.route('/staff/cashier')
@login_required
@role_required('cashier')
def cashier_dashboard():
    recent_payments = scoped(Payment).order_by(Payment.created_at.desc()).options(joinedload(Payment.student), joinedload(Payment.cashier)).limit(20).all()
    pending = scoped(Payment).filter_by(cleared=False).count()
    all_students = scoped(Student).order_by(Student.student_id).all()
    return render_template('staff/cashier/dashboard.html', payments=recent_payments, pending=pending, all_students=all_students)


@app.route('/staff/cashier/student/fees', methods=['GET', 'POST'])
@login_required
@role_required('cashier')
def cashier_student_fees():
    student = None
    fee_account = None
    payments = []
    detected_fee = None

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        student = scoped(Student).filter_by(student_id=student_id).first()
        if not student:
            flash('Student not found.', 'danger')
        else:
            fee_account = FeeAccount.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).first()
            payments = Payment.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).order_by(Payment.created_at.desc()).all()
            detected_fee = get_detected_fee(student)

    return render_template('staff/cashier/student_fees.html', student=student, fee_account=fee_account,
                           payments=payments, detected_fee=detected_fee)


@app.route('/staff/cashier/payment/add', methods=['POST'])
@login_required
@role_required('cashier')
def cashier_add_payment():
    student_id = request.form.get('student_id')
    amount = float(request.form.get('amount'))
    payment_method = request.form.get('payment_method')
    reference = request.form.get('reference')

    student = scoped(Student).filter_by(id=student_id).first()
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('cashier_dashboard'))

    payment = Payment(
        student_id=student.id, receipt_number=generate_receipt(),
        amount=amount, payment_date=datetime.now(),
        payment_method=payment_method, reference=reference,
        cashier_id=current_user.id, cleared=False,
        campus_id=current_user.campus_id
    )
    db.session.add(payment)

    fee_account = FeeAccount.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).first()
    if fee_account:
        fee_account.amount_paid += amount
        fee_account.balance = fee_account.total_fees - fee_account.amount_paid
    else:
        ensure_fee_account(student, amount_paid=amount)

    db.session.commit()
    log_activity('Payment recorded', f'${amount:.2f} for {student.full_name} ({student.reg_number})', user=current_user, student=student)
    flash(f'Payment recorded. Receipt: {payment.receipt_number}', 'success')
    return redirect(url_for('cashier_student_fees'))


@app.route('/staff/cashier/payment/clear/<int:payment_id>')
@login_required
@role_required('cashier')
def cashier_clear_payment(payment_id):
    payment = scoped_get_or_404(Payment, payment_id)
    payment.cleared = True
    payment.cleared_at = datetime.now()
    db.session.commit()
    flash(f'Payment {payment.receipt_number} cleared.', 'success')
    return redirect(request.referrer or url_for('cashier_dashboard'))


@app.route('/staff/cashier/setup-fees', methods=['POST'])
@login_required
@role_required('cashier')
def cashier_setup_fees():
    student_id = request.form.get('student_id')
    term = request.form.get('term')
    total_fees = float(request.form.get('total_fees'))

    student = scoped(Student).filter_by(id=student_id).first()
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('cashier_dashboard'))

    existing = FeeAccount.query.filter_by(student_id=student.id, term=term, campus_id=current_user.campus_id).first()
    if existing:
        existing.total_fees = total_fees
        existing.balance = total_fees - existing.amount_paid
    else:
        fee_account = FeeAccount(
            student_id=student.id, term=term, campus_id=current_user.campus_id,
            total_fees=total_fees, amount_paid=0.0, balance=total_fees
        )
        db.session.add(fee_account)

    db.session.commit()
    flash('Fee account updated.', 'success')
    return redirect(url_for('cashier_student_fees'))


# ============ ADMIN ROUTES ============

@app.route('/staff/admin')
@login_required
@role_required('admin')
def admin_dashboard():
    students = scoped(Student).order_by(Student.created_at.desc()).all()
    active_students = [s for s in students if s.is_active]
    inactive_students = [s for s in students if not s.is_active]
    return render_template('staff/admin/dashboard.html', students=students, active_students=active_students,
                           inactive_students=inactive_students)


@app.route('/staff/admin/student/add', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_add_student():
    subjects = scoped(Subject).all()
    o_level = [s for s in subjects if s.level == 'O']
    a_level = [s for s in subjects if s.level == 'A']
    if request.method == 'POST':
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        form = request.form.get('form')
        curriculum = request.form.get('curriculum')
        email = request.form.get('email')
        phone = request.form.get('phone')
        password = request.form.get('password')
        subject_ids = request.form.getlist('subjects')

        student_id = generate_student_id()
        reg_number = f'REG{student_id}'
        target_campus_id = get_selected_campus_id()

        student = Student(
            student_id=student_id, first_name=first_name, last_name=last_name,
            form=form, curriculum=curriculum, email=email or None, phone=phone, reg_number=reg_number,
            campus_id=target_campus_id
        )
        student.set_password(password or 'student123')

        db.session.add(student)
        db.session.flush()

        for sid in subject_ids:
            ss = StudentSubject(student_id=student.id, subject_id=int(sid))
            db.session.add(ss)

        term_fee = get_term_fee(form, campus_id=target_campus_id)
        if term_fee > 0:
            db.session.add(FeeAccount(student_id=student.id, term=get_current_term(), campus_id=target_campus_id,
                                      total_fees=term_fee, amount_paid=0.0, balance=term_fee))

        db.session.commit()
        log_activity('New student registered', f'{first_name} {last_name} ({reg_number})', user=current_user)
        flash(f'Student {first_name} {last_name} added. ID: {student_id}, Reg: {reg_number}', 'success')
        return redirect(url_for('admin_dashboard'))

    campuses = Campus.query.order_by(Campus.name).all() if getattr(g, 'is_super_admin', False) else []
    return render_template('staff/admin/add_student.html', subjects=subjects, o_level=o_level, a_level=a_level, campuses=campuses)


@app.route('/staff/admin/students/import', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'principal')
def admin_import_students():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('Please select an Excel file (.xlsx).', 'danger')
            return redirect(url_for('admin_import_students'))
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Unsupported file type. Please upload an .xlsx file.', 'danger')
            return redirect(url_for('admin_import_students'))
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            flash('Could not read the file. Ensure it is a valid Excel workbook.', 'danger')
            return redirect(url_for('admin_import_students'))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            flash('The file is empty.', 'danger')
            return redirect(url_for('admin_import_students'))

        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
        required = ('first_name', 'last_name', 'form')
        if not all(h in header for h in required):
            flash('Required columns: first_name, last_name, form. Optional: curriculum, email, phone.', 'danger')
            return redirect(url_for('admin_import_students'))
        idx = {name: header.index(name) for name in ('first_name', 'last_name', 'form', 'curriculum', 'email', 'phone') if name in header}

        target_campus_id = get_selected_campus_id()
        term = get_current_term()
        added = skipped = 0
        for i, row in enumerate(rows[1:], start=2):
            def get_val(name):
                if name not in idx or idx[name] >= len(row):
                    return None
                v = row[idx[name]]
                s = str(v).strip() if v is not None else ''
                return s if s and s.lower() != 'none' else None

            first_name = get_val('first_name')
            last_name = get_val('last_name')
            form = get_val('form')
            if not first_name or not last_name or not form:
                skipped += 1
                continue
            if form.isdigit():
                form = f'Form {form}'
            curriculum = (get_val('curriculum') or 'ZIMSEC').upper()
            if curriculum not in ('ZIMSEC', 'CAMBRIDGE'):
                curriculum = 'ZIMSEC'
            email = get_val('email')
            phone = get_val('phone')

            student_id = generate_student_id()
            student = Student(student_id=student_id, first_name=first_name, last_name=last_name,
                              form=form, curriculum=curriculum, email=email, phone=phone,
                              reg_number=f'REG{student_id}', campus_id=target_campus_id)
            student.set_password('student123')
            db.session.add(student)
            db.session.flush()
            term_fee = get_term_fee(form, campus_id=target_campus_id)
            if term_fee > 0:
                db.session.add(FeeAccount(student_id=student.id, term=term, campus_id=target_campus_id,
                                          total_fees=term_fee, amount_paid=0.0, balance=term_fee))
            added += 1

        db.session.commit()
        log_activity('Bulk students imported', f'{added} students added via Excel upload', user=current_user)
        flash(f'Import complete: {added} student(s) added, {skipped} row(s) skipped. Default password: student123', 'success' if added else 'warning')
        if current_user.role == 'principal':
            return redirect(url_for('principal_dashboard'))
        return redirect(url_for('admin_dashboard'))

    campuses = Campus.query.order_by(Campus.name).all() if getattr(g, 'is_super_admin', False) else []
    return render_template('staff/admin/import_students.html', campuses=campuses)


@app.route('/staff/admin/students/import/template')
@login_required
@role_required('admin', 'principal')
def admin_import_student_template():
    import io
    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.append(['first_name', 'last_name', 'form', 'curriculum', 'email', 'phone'])
    ws.append(['Tendai', 'Moyo', 'Form 1', 'ZIMSEC', 'tendai@example.com', '0771111111'])
    ws.append(['Chipo', 'Ncube', '2', 'CAMBRIDGE', 'chipo@example.com', '0772222222'])
    for col, width in zip('ABCDEF', [16, 16, 10, 12, 26, 14]):
        ws.column_dimensions[col].width = width
    wb.save(buf)
    buf.seek(0)
    return app.response_class(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              headers={'Content-Disposition': 'attachment; filename=student_template.xlsx'})


@app.route('/staff/admin/student/edit/<int:student_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_student(student_id):
    student = scoped_get_or_404(Student, student_id)
    original_form = student.form
    subjects = scoped(Subject).all()
    enrolled_ids = [ss.subject_id for ss in student.subjects]
    o_level = [s for s in subjects if s.level == 'O']
    a_level = [s for s in subjects if s.level == 'A']

    if request.method == 'POST':
        student.first_name = request.form.get('first_name')
        student.last_name = request.form.get('last_name')
        student.form = request.form.get('form')
        student.curriculum = request.form.get('curriculum')
        student.email = request.form.get('email') or None
        student.phone = request.form.get('phone')
        if request.form.get('password'):
            student.set_password(request.form.get('password'))

        StudentSubject.query.filter_by(student_id=student.id).delete()
        for sid in request.form.getlist('subjects'):
            ss = StudentSubject(student_id=student.id, subject_id=int(sid))
            db.session.add(ss)

        if student.form != original_form:
            fee_account = FeeAccount.query.filter_by(student_id=student.id, term=get_current_term()).first()
            if fee_account:
                new_fee = get_term_fee(student.form, campus_id=student.campus_id)
                if new_fee > 0:
                    fee_account.total_fees = new_fee
                    fee_account.balance = new_fee - fee_account.amount_paid

        db.session.commit()
        flash('Student updated.', 'success')
        return redirect(url_for('admin_dashboard'))

    return render_template('staff/admin/edit_student.html', student=student, subjects=subjects, enrolled_ids=enrolled_ids, o_level=o_level, a_level=a_level)


@app.route('/staff/admin/student/deactivate/<int:student_id>')
@login_required
@role_required('admin')
def admin_deactivate_student(student_id):
    student = scoped_get_or_404(Student, student_id)
    student.is_active = not student.is_active
    status = 'reactivated' if student.is_active else 'deactivated (transferred)'
    db.session.commit()
    flash(f'Student {student.full_name} {status}.', 'info')
    return redirect(url_for('admin_dashboard'))


@app.route('/staff/admin/student/remove/<int:student_id>')
@app.route('/staff/principal/student/remove/<int:student_id>')
@login_required
@role_required('admin', 'principal')
def remove_student(student_id):
    student = scoped_get_or_404(Student, student_id)
    name = student.full_name
    sid_label = student.student_id
    sid = student.id

    for model in (StudentSubject, MonthlyTest, ExamMark, PrincipalComment, FeeAccount, Payment, TeacherRemark):
        db.session.query(model).filter(model.student_id == sid).delete(synchronize_session=False)
    db.session.query(ActivityLog).filter(ActivityLog.student_id == sid).update(
        {ActivityLog.student_id: None}, synchronize_session=False)

    db.session.delete(student)
    db.session.commit()
    log_activity('Student removed', f'{name} ({sid_label}) permanently deleted', user=current_user)
    flash(f'Student {name} has been permanently removed.', 'warning')
    if current_user.role == 'principal':
        return redirect(url_for('principal_dashboard'))
    return redirect(url_for('admin_dashboard'))


@app.route('/staff/admin/student/<int:student_id>/remove-subjects', methods=['POST'])
@login_required
@role_required('admin')
def admin_remove_student_subjects(student_id):
    student = scoped_get_or_404(Student, student_id)
    count = StudentSubject.query.filter_by(student_id=student.id).count()
    StudentSubject.query.filter_by(student_id=student.id).delete()
    db.session.commit()
    log_activity('Student removed from all subjects', f'{student.full_name} ({count} subjects)', user=current_user)
    flash(f'Removed {student.full_name} from all {count} subject(s).', 'success')
    return redirect(url_for('admin_edit_student', student_id=student.id))


@app.route('/staff/admin/activity/add', methods=['POST'])
@login_required
@role_required('admin')
def admin_add_activity():
    action = request.form.get('action')
    description = request.form.get('description')
    visibility = request.form.get('visibility', 'public')
    if action:
        log_activity(action, description, user=current_user, visibility=visibility)
        flash('Activity logged.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/staff/admin/activities')
@login_required
@role_required('admin')
def admin_activities():
    activities_list = scoped(Activity).order_by(Activity.created_at.desc()).all()
    return render_template('staff/admin/activities.html', activities_list=activities_list)


@app.route('/staff/admin/activities/add', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_add_new_activity():
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        visibility = request.form.get('visibility', 'all')
        if title:
            activity = Activity(title=title, description=description,
                                visibility=visibility, created_by=current_user.id,
                                campus_id=current_user.campus_id)
            db.session.add(activity)
            db.session.commit()
            flash('Activity created.', 'success')
            return redirect(url_for('admin_activities'))
        flash('Title is required.', 'danger')
    return render_template('staff/admin/add_activity.html')


@app.route('/staff/admin/activities/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_activity(id):
    activity = scoped_get_or_404(Activity, id)
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        visibility = request.form.get('visibility', 'all')
        if title:
            activity.title = title
            activity.description = description
            activity.visibility = visibility
            db.session.commit()
            flash('Activity updated.', 'success')
            return redirect(url_for('admin_activities'))
        flash('Title is required.', 'danger')
    return render_template('staff/admin/edit_activity.html', activity=activity)


@app.route('/staff/admin/activities/delete/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_activity(id):
    activity = scoped_get_or_404(Activity, id)
    db.session.delete(activity)
    db.session.commit()
    flash('Activity deleted.', 'success')
    return redirect(url_for('admin_activities'))


# ============ PRINCIPAL ROUTES ============

@app.route('/staff/principal')
@login_required
@role_required('principal')
def principal_dashboard():
    students = scoped(Student).order_by(Student.student_id).all()
    return render_template('staff/principal/dashboard.html', students=students)


@app.route('/staff/principal/student/<int:student_id>/results')
@login_required
@role_required('principal')
def principal_student_results(student_id):
    student = scoped_get_or_404(Student, student_id)
    subjects = [ss.subject for ss in student.subjects if ss.subject]
    monthly_tests = MonthlyTest.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).order_by(MonthlyTest.academic_year.desc(), MonthlyTest.term).all()
    exam_marks_list = ExamMark.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).order_by(ExamMark.academic_year.desc(), ExamMark.term).all()
    principal_comments = PrincipalComment.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).all()
    pc_by_key = {}
    for pc in principal_comments:
        pc_by_key[(pc.subject_id, pc.term, pc.academic_year)] = pc
    return render_template('staff/principal/student_results.html', student=student, subjects=subjects,
                         monthly_tests=monthly_tests, exam_marks=exam_marks_list, pc_by_key=pc_by_key)


@app.route('/staff/principal/comment/save', methods=['POST'])
@login_required
@role_required('principal')
def principal_save_comment():
    student_id = request.form.get('student_id')
    subject_id = request.form.get('subject_id')
    term = request.form.get('term')
    academic_year = request.form.get('academic_year')
    comment = request.form.get('comment')

    existing = PrincipalComment.query.filter_by(
        student_id=student_id, subject_id=subject_id,
        term=term, academic_year=academic_year,
        campus_id=current_user.campus_id
    ).first()
    if existing:
        existing.comment = comment
    else:
        pc = PrincipalComment(
            student_id=student_id, subject_id=subject_id,
            term=term, academic_year=academic_year, comment=comment,
            campus_id=current_user.campus_id
        )
        db.session.add(pc)
    db.session.commit()
    flash('Comment saved.', 'success')
    return redirect(url_for('principal_student_results', student_id=student_id))


@app.route('/staff/principal/staff/add', methods=['GET', 'POST'])
@login_required
@role_required('principal')
def principal_add_staff():
    subjects = scoped(Subject).all()
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        subject_id = request.form.get('subject_id')

        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('principal_add_staff'))

        errors = password_strength_errors(password)
        if errors:
            flash(errors[0], 'danger')
            return redirect(url_for('principal_add_staff'))

        reg_number = f'STF{username.upper()}'
        target_campus_id = get_selected_campus_id()

        on_leave = request.form.get('on_leave') == 'on'
        leave_start = request.form.get('leave_start')
        leave_end = request.form.get('leave_end')
        if on_leave:
            if not leave_start or not leave_end:
                flash('Please provide leave dates when marking staff as on leave.', 'danger')
                return redirect(url_for('principal_add_staff'))
            if leave_start > leave_end:
                flash('Leave start date cannot be after end date.', 'danger')
                return redirect(url_for('principal_add_staff'))

        if role == 'teacher' and subject_id:
            try:
                subject_id = int(subject_id)
            except (TypeError, ValueError):
                flash('Please select a valid subject for the teacher.', 'danger')
                return redirect(url_for('principal_add_staff'))
        final_email = staff_email(username, email)
        if User.query.filter(User.email == final_email).first():
            flash('That email is already used by another staff member.', 'danger')
            return redirect(url_for('principal_add_staff'))

        user = User(
            username=username, email=final_email, role=role, full_name=full_name,
            phone=phone, reg_number=reg_number,
            subject_id=subject_id if role == 'teacher' else None,
            campus_id=target_campus_id
        )
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        if on_leave:
            db.session.add(StaffLeave(
                campus_id=target_campus_id, user_id=user.id,
                leave_type=request.form.get('leave_type') or 'Annual',
                start_date=datetime.strptime(leave_start, '%Y-%m-%d').date(),
                end_date=datetime.strptime(leave_end, '%Y-%m-%d').date(),
                status='Approved', reason=request.form.get('leave_reason') or None
            ))

        db.session.commit()
        flash(f'Staff {full_name} added as {role}. Reg: {reg_number}', 'success')
        return redirect(url_for('principal_dashboard'))

    campuses = Campus.query.order_by(Campus.name).all() if getattr(g, 'is_super_admin', False) else []
    return render_template('staff/principal/add_staff.html', subjects=subjects, campuses=campuses)


@app.route('/staff/principal/staff/edit/<int:staff_id>', methods=['GET', 'POST'])
@login_required
@role_required('principal')
def principal_edit_staff(staff_id):
    staff_member = scoped_get_or_404(User, staff_id)
    subjects = scoped(Subject).all()

    if request.method == 'POST':
        final_email = staff_email(staff_member.username, request.form.get('email'))
        if User.query.filter(User.email == final_email, User.id != staff_member.id).first():
            flash('That email is already used by another staff member.', 'danger')
            return redirect(url_for('principal_edit_staff', staff_id=staff_id))
        staff_member.full_name = request.form.get('full_name')
        staff_member.email = final_email
        staff_member.phone = request.form.get('phone')
        staff_member.role = request.form.get('role')
        subj = request.form.get('subject_id')
        if subj:
            try:
                subj = int(subj)
            except (TypeError, ValueError):
                flash('Please select a valid subject for the teacher.', 'danger')
                return redirect(url_for('principal_edit_staff', staff_id=staff_id))
        staff_member.subject_id = subj if request.form.get('role') == 'teacher' else None
        if request.form.get('password'):
            staff_member.set_password(request.form.get('password'))
        db.session.commit()
        flash('Staff updated.', 'success')
        return redirect(url_for('principal_dashboard'))

    return render_template('staff/principal/edit_staff.html', staff=staff_member, subjects=subjects)


@app.route('/staff/principal/staff/fire/<int:staff_id>')
@login_required
@role_required('principal')
def principal_fire_staff(staff_id):
    staff_member = scoped_get_or_404(User, staff_id)
    if staff_member.role == 'principal':
        flash('Cannot fire the principal.', 'danger')
        return redirect(url_for('principal_dashboard'))
    full_name = staff_member.full_name
    db.session.delete(staff_member)
    db.session.commit()
    flash(f'{full_name} has been fired and login removed.', 'warning')
    return redirect(url_for('principal_dashboard'))


@app.route('/staff/principal/student/toggle-status/<int:student_id>')
@login_required
@role_required('principal')
def principal_toggle_student_status(student_id):
    student = scoped_get_or_404(Student, student_id)
    student.is_active = not student.is_active
    status = 'reactivated' if student.is_active else 'deactivated'
    db.session.commit()
    flash(f'Student {student.full_name} has been {status}.', 'warning')
    return redirect(url_for('principal_dashboard'))


@app.route('/staff/principal/staff/leave', methods=['GET', 'POST'])
@login_required
@role_required('principal')
def principal_staff_leave():
    cid = current_user.campus_id
    staff_list = scoped(User).filter(User.role.notin_(['principal', 'super_admin'])).order_by(User.full_name).all()
    now_date = date.today()

    if request.method == 'POST':
        staff_id = request.form.get('staff_id')
        leave_type = request.form.get('leave_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        reason = request.form.get('reason')
        if not staff_id or not leave_type or not start_date or not end_date:
            flash('Please fill in all required fields.', 'danger')
            return redirect(url_for('principal_staff_leave'))
        staff_member = scoped_get_or_404(User, int(staff_id))
        if start_date > end_date:
            flash('Start date cannot be after end date.', 'danger')
            return redirect(url_for('principal_staff_leave'))
        leave = StaffLeave(
            campus_id=cid, user_id=staff_member.id, leave_type=leave_type,
            start_date=datetime.strptime(start_date, '%Y-%m-%d').date(),
            end_date=datetime.strptime(end_date, '%Y-%m-%d').date(),
            status='Approved', reason=reason or None
        )
        db.session.add(leave)
        db.session.commit()
        log_activity('Staff leave recorded',
                     f'{staff_member.full_name} on {leave_type} leave '
                     f'({start_date} to {end_date})', user=current_user, campus_id=cid)
        flash(f'Leave recorded for {staff_member.full_name}.', 'success')
        return redirect(url_for('principal_staff_leave'))

    leaves = scoped(StaffLeave).order_by(StaffLeave.created_at.desc()).options(joinedload(StaffLeave.staff)).all()
    on_leave_ids = {l.user_id for l in leaves
                    if l.status == 'Approved' and l.start_date <= now_date <= l.end_date}
    return render_template('staff/principal/staff_leave.html', staff_list=staff_list,
                           leaves=leaves, on_leave_ids=on_leave_ids, now_date=now_date)


@app.route('/staff/principal/staff/leave/<int:leave_id>/<string:action>', methods=['POST'])
@login_required
@role_required('principal')
def principal_leave_action(leave_id, action):
    leave = scoped_get_or_404(StaffLeave, leave_id)
    if action == 'approve':
        leave.status = 'Approved'
        flash(f'{leave.staff.full_name}\'s leave approved.', 'success')
    elif action == 'reject':
        leave.status = 'Rejected'
        flash(f'{leave.staff.full_name}\'s leave rejected.', 'warning')
    elif action == 'delete':
        name = leave.staff.full_name
        db.session.delete(leave)
        db.session.commit()
        flash(f'Leave record for {name} deleted.', 'info')
        return redirect(url_for('principal_staff_leave'))
    else:
        flash('Unknown action.', 'danger')
        return redirect(url_for('principal_staff_leave'))
    db.session.commit()
    log_activity(f'Staff leave {leave.status.lower()}',
                 f'{leave.staff.full_name} ({leave.leave_type})', user=current_user,
                 campus_id=current_user.campus_id)
    return redirect(url_for('principal_staff_leave'))


def _block_super_admin():
    """Timetable management is reserved for campus principals."""
    if getattr(current_user, 'role', '') == 'super_admin':
        flash('Timetable management is not available for super admin.', 'danger')
        return redirect(url_for('super_admin_dashboard'))
    return None


@app.route('/staff/principal/timetables')
@login_required
@role_required('principal')
def principal_timetables():
    blocked = _block_super_admin()
    if blocked:
        return blocked
    subjects = scoped(Subject).all()
    teachers = scoped(User).filter_by(role='teacher').all()
    timetables = scoped(Timetable).order_by(Timetable.form, Timetable.day_of_week, Timetable.start_time).all()
    exam_timetables = scoped(ExamTimetable).order_by(ExamTimetable.exam_date, ExamTimetable.start_time).all()
    return render_template('staff/principal/timetables.html', subjects=subjects, teachers=teachers, timetables=timetables, exam_timetables=exam_timetables)


@app.route('/staff/principal/timetable/add', methods=['POST'])
@login_required
@role_required('principal')
def principal_add_timetable():
    blocked = _block_super_admin()
    if blocked:
        return blocked
    tt = Timetable(
        form=request.form.get('form'), day_of_week=request.form.get('day_of_week'),
        subject_id=int(request.form.get('subject_id')), start_time=request.form.get('start_time'),
        end_time=request.form.get('end_time'), teacher_id=int(request.form.get('teacher_id')),
        room=request.form.get('room'), campus_id=current_user.campus_id
    )
    db.session.add(tt)
    db.session.commit()
    flash('Timetable entry added.', 'success')
    return redirect(url_for('principal_timetables'))


@app.route('/staff/principal/exam-timetable/add', methods=['POST'])
@login_required
@role_required('principal')
def principal_add_exam_timetable():
    blocked = _block_super_admin()
    if blocked:
        return blocked
    ett = ExamTimetable(
        form=request.form.get('form'), subject_id=int(request.form.get('subject_id')),
        exam_date=datetime.strptime(request.form.get('exam_date'), '%Y-%m-%d').date(),
        start_time=request.form.get('start_time'), end_time=request.form.get('end_time'),
        room=request.form.get('room'), campus_id=current_user.campus_id
    )
    db.session.add(ett)
    db.session.commit()
    flash('Exam timetable entry added.', 'success')
    return redirect(url_for('principal_timetables'))


@app.route('/staff/principal/timetable/delete/<int:tt_id>')
@login_required
@role_required('principal')
def principal_delete_timetable(tt_id):
    blocked = _block_super_admin()
    if blocked:
        return blocked
    db.session.delete(scoped_get_or_404(Timetable, tt_id))
    db.session.commit()
    flash('Timetable entry deleted.', 'success')
    return redirect(url_for('principal_timetables'))


@app.route('/staff/principal/exam-timetable/delete/<int:ett_id>')
@login_required
@role_required('principal')
def principal_delete_exam_timetable(ett_id):
    blocked = _block_super_admin()
    if blocked:
        return blocked
    db.session.delete(scoped_get_or_404(ExamTimetable, ett_id))
    db.session.commit()
    flash('Exam timetable entry deleted.', 'success')
    return redirect(url_for('principal_timetables'))


@app.route('/staff/principal/fee-settings', methods=['GET', 'POST'])
@login_required
@role_required('principal')
def principal_fee_settings():
    if request.method == 'POST':
        form = request.form.get('form', '').strip()
        term_fee = float(request.form.get('term_fee', 0))
        if form and term_fee > 0:
            existing = FeeSetting.query.filter_by(form=form, campus_id=current_user.campus_id).first()
            if existing:
                existing.term_fee = term_fee
            else:
                db.session.add(FeeSetting(form=form, term_fee=term_fee, campus_id=current_user.campus_id))
            db.session.commit()
            flash(f'Fee for {form} set to ${term_fee:.2f}.', 'success')
        return redirect(url_for('principal_fee_settings'))

    fee_settings = scoped(FeeSetting).order_by(FeeSetting.form).all()
    forms = ['Form 1', 'Form 2', 'Form 3', 'Form 4', 'Form 5', 'Form 6']
    return render_template('staff/principal/fee_settings.html', fee_settings=fee_settings, forms=forms)


@app.route('/staff/principal/fee-settings/delete/<int:fs_id>')
@login_required
@role_required('principal')
def principal_delete_fee_setting(fs_id):
    fs = scoped_get_or_404(FeeSetting, fs_id)
    db.session.delete(fs)
    db.session.commit()
    flash('Fee setting removed.', 'success')
    return redirect(url_for('principal_fee_settings'))


@app.route('/staff/principal/student/<int:student_id>/edit-fees', methods=['GET', 'POST'])
@login_required
@role_required('principal')
def principal_edit_student_fees(student_id):
    student = scoped_get_or_404(Student, student_id)
    fee_account = FeeAccount.query.filter_by(student_id=student.id, campus_id=current_user.campus_id).first()
    default_fee = fee_account.total_fees if fee_account else get_term_fee(student.form, campus_id=current_user.campus_id)

    if request.method == 'POST':
        total_fees = float(request.form.get('total_fees', 0))
        amount_paid = float(request.form.get('amount_paid', 0))
        if not fee_account:
            fee_account = FeeAccount(student_id=student.id, term=get_current_term(), campus_id=current_user.campus_id)
            db.session.add(fee_account)
        fee_account.total_fees = total_fees
        fee_account.amount_paid = amount_paid
        fee_account.balance = total_fees - amount_paid
        db.session.commit()
        flash(f'Fees updated for {student.full_name}.', 'success')
        return redirect(url_for('principal_dashboard'))

    return render_template('staff/principal/edit_fees.html', student=student, fee_account=fee_account, default_fee=default_fee)


# ============ EXPORT ROUTES ============

@app.route('/staff/principal/export/zimsec-candidates')
@login_required
@role_required('principal')
def export_zimsec_candidates():
    students = scoped(Student).filter_by(curriculum='ZIMSEC', is_active=True).order_by(Student.form, Student.last_name).all()
    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Student ID', 'Reg Number', 'Full Name', 'Form', 'Subjects', 'Gender'])
    for s in students:
        subs = ', '.join(ss.subject.name for ss in s.subjects)
        writer.writerow([s.student_id, s.reg_number, s.full_name, s.form, subs, ''])
    response = app.response_class(output.getvalue(), mimetype='text/csv',
                                  headers={'Content-Disposition': 'attachment; filename=zimsec_candidates.csv'})
    log_activity('Exported ZIMSEC candidate list', user=current_user)
    return response


@app.route('/staff/principal/export/ministry-report')
@login_required
@role_required('principal')
def export_ministry_report():
    term = 'Term 1'
    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Form', 'Total Students', 'Total Passed', 'Pass Rate', 'Total Expected Fees', 'Total Collected', 'Outstanding'])
    forms = ['Form 1', 'Form 2', 'Form 3', 'Form 4', 'Form 5', 'Form 6']
    all_students = scoped(Student).all()
    student_form = {s.id: s.form for s in all_students}
    all_exams = ExamMark.query.filter_by(term=term, campus_id=current_user.campus_id).all()
    by_form = {}
    for e in all_exams:
        f = student_form.get(e.student_id)
        if f:
            by_form.setdefault(f, []).append(e)
    fee_total = compute_expected_fees(term, current_user.campus_id)
    fee_collected = db.session.query(db.func.sum(Payment.amount)).filter(Payment.cleared == True, Payment.campus_id == current_user.campus_id).scalar() or 0
    outstanding_fees = fee_total - fee_collected
    for f in forms:
        total = sum(1 for s in all_students if s.form == f)
        exams = by_form.get(f, [])
        passed = sum(1 for e in exams if e.total_marks > 0 and e.marks / e.total_marks * 100 >= 50)
        pass_rate = round(passed / total * 100, 1) if total > 0 else 0
        writer.writerow([f, total, passed, f'{pass_rate}%', f'${fee_total:.2f}', f'${fee_collected:.2f}', f'${outstanding_fees:.2f}'])
    response = app.response_class(output.getvalue(), mimetype='text/csv',
                                  headers={'Content-Disposition': 'attachment; filename=ministry_report.csv'})
    log_activity('Exported Ministry report', user=current_user)
    return response


# ============ STUDENT ROUTES ============

@app.route('/student/dashboard')
@login_required
@student_required
def student_dashboard():
    subjects = [ss.subject for ss in current_user.subjects if ss.subject]
    monthly_tests = MonthlyTest.query.filter_by(student_id=current_user.id).all()
    exam_marks_list = ExamMark.query.filter_by(student_id=current_user.id).all()

    total_exams = len(exam_marks_list)
    passed_exams = sum(1 for e in exam_marks_list if e.total_marks > 0 and (e.marks / e.total_marks * 100) >= 50)
    pass_rate = round(passed_exams / total_exams * 100) if total_exams else 0

    subject_pass_rates = []
    for subj in subjects:
        marks = [e for e in exam_marks_list if e.subject_id == subj.id]
        if marks:
            passed = sum(1 for e in marks if e.total_marks > 0 and (e.marks / e.total_marks * 100) >= 50)
            subject_pass_rates.append({'subject': subj.name, 'rate': round(passed / len(marks) * 100),
                                       'passed': passed, 'total': len(marks)})

    fee_account = FeeAccount.query.filter_by(student_id=current_user.id).first()
    payments = Payment.query.filter_by(student_id=current_user.id).order_by(Payment.created_at.desc()).limit(10).all()
    timetables = Timetable.query.filter_by(form=current_user.form, campus_id=current_user.campus_id).order_by(Timetable.day_of_week, Timetable.start_time).all()
    exam_timetables = ExamTimetable.query.filter_by(form=current_user.form, campus_id=current_user.campus_id).order_by(ExamTimetable.exam_date).all()

    # Get teacher for each subject
    subject_teachers = {}
    for subj in subjects:
        teacher = User.query.filter_by(subject_id=subj.id, role='teacher').first()
        subject_teachers[subj.id] = teacher

    # Get teacher remarks for this student
    teacher_remarks = TeacherRemark.query.filter_by(student_id=current_user.id).order_by(TeacherRemark.created_at.desc()).all()
    recent_activity_posts = scoped(Activity).filter_by(visibility='all').order_by(Activity.created_at.desc()).limit(10).all()

    return render_template('student/dashboard.html', subjects=subjects, monthly_tests=monthly_tests,
                         exam_marks=exam_marks_list, fee_account=fee_account, payments=payments,
                         timetables=timetables, exam_timetables=exam_timetables, subject_teachers=subject_teachers,
                         teacher_remarks=teacher_remarks, recent_activity_posts=recent_activity_posts,
                         pass_rate=pass_rate, subject_pass_rates=subject_pass_rates)


@app.route('/student/results')
@login_required
@student_required
def student_results():
    fee_account = FeeAccount.query.filter_by(student_id=current_user.id).first()
    fees_cleared = not fee_account or fee_account.balance <= 0

    subjects = [ss.subject for ss in current_user.subjects if ss.subject]
    monthly_tests = MonthlyTest.query.filter_by(student_id=current_user.id).all()
    exam_marks_list = ExamMark.query.filter_by(student_id=current_user.id).all()

    total_exams = len(exam_marks_list)
    passed_exams = sum(1 for e in exam_marks_list if e.total_marks > 0 and (e.marks / e.total_marks * 100) >= 50)
    pass_rate = round(passed_exams / total_exams * 100) if total_exams else 0

    subject_pass_rates = []
    for subj in subjects:
        marks = [e for e in exam_marks_list if e.subject_id == subj.id]
        if marks:
            passed = sum(1 for e in marks if e.total_marks > 0 and (e.marks / e.total_marks * 100) >= 50)
            subject_pass_rates.append({'subject': subj.name, 'rate': round(passed / len(marks) * 100),
                                       'passed': passed, 'total': len(marks)})

    principal_comments = PrincipalComment.query.filter_by(student_id=current_user.id).all()
    pc_by_key = {}
    for pc in principal_comments:
        pc_by_key[(pc.subject_id, pc.term, pc.academic_year)] = pc

    results = {}
    prev_monthly_all = MonthlyTest.query.filter(
        MonthlyTest.student_id == current_user.id,
        MonthlyTest.academic_year < str(datetime.now().year)
    ).order_by(MonthlyTest.academic_year, MonthlyTest.term).all()
    prev_exam_all = ExamMark.query.filter(
        ExamMark.student_id == current_user.id,
        ExamMark.academic_year < str(datetime.now().year)
    ).order_by(ExamMark.academic_year, ExamMark.term).all()

    prev_monthly_by_subj = {}
    for mt in prev_monthly_all:
        prev_monthly_by_subj.setdefault(mt.subject_id, []).append(mt)
    prev_exam_by_subj = {}
    for em in prev_exam_all:
        prev_exam_by_subj.setdefault(em.subject_id, []).append(em)

    for subj in subjects:
        s_monthly = [mt for mt in monthly_tests if mt.subject_id == subj.id]
        s_exam = [em for em in exam_marks_list if em.subject_id == subj.id]

        level = 'A' if '6' in current_user.form or '5' in current_user.form else 'O'
        prev_monthly = prev_monthly_by_subj.get(subj.id, [])
        prev_exam = prev_exam_by_subj.get(subj.id, [])

        pc = pc_by_key.get((subj.id, 'Term 1', str(datetime.now().year)))

        results[subj.name] = {
            'monthly': s_monthly, 'exam': s_exam,
            'prev_monthly': prev_monthly, 'prev_exam': prev_exam,
            'level': level, 'principal_comment': pc.comment if pc else None
        }

    return render_template('student/results.html', results=results, fees_cleared=fees_cleared, fee_account=fee_account,
                         pass_rate=pass_rate, subject_pass_rates=subject_pass_rates)


@app.route('/student/finances')
@login_required
@student_required
def student_finances():
    fee_account = FeeAccount.query.filter_by(student_id=current_user.id).first()
    payments = Payment.query.filter_by(student_id=current_user.id).order_by(Payment.created_at.desc()).all()
    return render_template('student/finances.html', fee_account=fee_account, payments=payments)


@app.route('/student/timetables')
@login_required
@student_required
def student_timetables():
    timetables = Timetable.query.filter_by(form=current_user.form, campus_id=current_user.campus_id).order_by(Timetable.day_of_week, Timetable.start_time).all()
    exam_timetables_list = ExamTimetable.query.filter_by(form=current_user.form, campus_id=current_user.campus_id).order_by(ExamTimetable.exam_date).all()

    timetable_by_day = {}
    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']:
        timetable_by_day[day] = [tt for tt in timetables if tt.day_of_week == day]

    return render_template('student/timetables.html', timetable_by_day=timetable_by_day, exam_timetables=exam_timetables_list)


@app.route('/student/subjects')
@login_required
@student_required
def student_subjects():
    subjects = [ss.subject for ss in current_user.subjects if ss.subject]
    subject_teachers = {}
    for subj in subjects:
        teacher = User.query.filter_by(subject_id=subj.id, role='teacher').first()
        subject_teachers[subj.id] = teacher
    return render_template('student/subjects.html', subjects=subjects, subject_teachers=subject_teachers)


# ============ SUPER ADMIN ROUTES ============

@app.route('/super-admin/dashboard')
@login_required
@super_admin_required
def super_admin_dashboard():
    campuses = Campus.query.order_by(Campus.name).all()
    total_campuses = len(campuses)
    total_students = Student.query.count()
    total_staff = User.query.filter(User.role != 'super_admin').count()
    total_payments = db.session.query(db.func.sum(Payment.amount)).filter(Payment.cleared == True).scalar() or 0
    return render_template('super_admin/dashboard.html', campuses=campuses, total_campuses=total_campuses,
                           total_students=total_students, total_staff=total_staff, total_payments=total_payments)


@app.route('/super-admin/campuses')
@login_required
@super_admin_required
def super_admin_campuses():
    campuses = Campus.query.order_by(Campus.name).all()
    return render_template('super_admin/campuses.html', campuses=campuses)


@app.route('/super-admin/campuses/add', methods=['POST'])
@login_required
@super_admin_required
def super_admin_add_campus():
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip().upper()
    address = request.form.get('address', '').strip()
    if not name or not code:
        flash('Campus name and code are required.', 'danger')
        return redirect(url_for('super_admin_campuses'))
    if Campus.query.filter_by(code=code).first():
        flash(f'Campus code {code} already exists.', 'danger')
        return redirect(url_for('super_admin_campuses'))
    db.session.add(Campus(name=name, code=code, address=address))
    db.session.commit()
    new_campus = Campus.query.filter_by(code=code).first()
    if new_campus is not None:
        seed_campus_subjects(new_campus.id)
    log_activity('Campus created', f'{name} ({code})', user=current_user)
    flash(f'Campus {name} ({code}) created.', 'success')
    return redirect(url_for('super_admin_campuses'))


@app.route('/super-admin/campuses/<int:campus_id>/edit', methods=['GET', 'POST'])
@login_required
@super_admin_required
def super_admin_edit_campus(campus_id):
    campus = Campus.query.get_or_404(campus_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip().upper()
        address = request.form.get('address', '').strip()
        if not name or not code:
            flash('Campus name and code are required.', 'danger')
            return redirect(url_for('super_admin_edit_campus', campus_id=campus.id))
        clash = Campus.query.filter(Campus.code == code, Campus.id != campus.id).first()
        if clash:
            flash(f'Campus code {code} is already used by {clash.name}.', 'danger')
            return redirect(url_for('super_admin_edit_campus', campus_id=campus.id))
        old = f'{campus.name} ({campus.code})'
        campus.name = name
        campus.code = code
        campus.address = address
        db.session.commit()
        log_activity('Campus updated', f'{old} -> {name} ({code})', user=current_user)
        flash(f'Campus {name} updated.', 'success')
        return redirect(url_for('super_admin_campuses'))
    return render_template('super_admin/edit_campus.html', campus=campus)



@app.route('/super-admin/campuses/<int:campus_id>/staff', methods=['GET', 'POST'])
@login_required
@super_admin_required
def super_admin_campus_staff(campus_id):
    campus = Campus.query.get_or_404(campus_id)
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        subject_id = request.form.get('subject_id')
        if not username or not full_name:
            flash('Username and full name are required.', 'danger')
            return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))
        errors = password_strength_errors(password)
        if errors:
            flash(errors[0], 'danger')
            return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))
        if role == 'teacher' and not subject_id:
            flash('Please select the subject the teacher teaches.', 'danger')
            return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))
        if role == 'teacher':
            try:
                subject_id = int(subject_id)
            except (TypeError, ValueError):
                flash('Please select a valid subject for the teacher.', 'danger')
                return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))
        final_email = staff_email(username, email)
        if User.query.filter(User.email == final_email).first():
            flash('That email is already used by another staff member.', 'danger')
            return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))
        user = User(username=username, email=final_email, role=role, full_name=full_name,
                    phone=phone or None, reg_number=f'STF{username.upper()}', campus_id=campus.id,
                    subject_id=subject_id if role == 'teacher' else None)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        log_activity('Staff assigned to campus', f'{full_name} ({role}) -> {campus.name}', user=current_user)
        flash(f'{full_name} added as {role} at {campus.name}.', 'success')
        return redirect(url_for('super_admin_campus_staff', campus_id=campus.id))

    staff = User.query.filter(User.campus_id == campus.id, User.role != 'super_admin').order_by(User.full_name).all()
    if not seed_campus_subjects(campus.id):
        flash('Could not seed standard subjects for this campus; the page may be incomplete.', 'warning')
    subjects = Subject.query.filter_by(campus_id=campus.id).order_by(Subject.name).all()
    return render_template('super_admin/campus_staff.html', campus=campus, staff=staff, subjects=subjects)


@app.route('/super-admin/campuses/<int:campus_id>/staff/fire/<int:staff_id>', methods=['POST'])
@login_required
@super_admin_required
def super_admin_fire_staff(campus_id, staff_id):
    staff_member = User.query.get_or_404(staff_id)
    if staff_member.campus_id != campus_id:
        flash('Staff member not in this campus.', 'danger')
        return redirect(url_for('super_admin_campus_staff', campus_id=campus_id))
    if staff_member.role == 'super_admin':
        flash('Cannot remove a super admin.', 'danger')
        return redirect(url_for('super_admin_campus_staff', campus_id=campus_id))
    full_name = staff_member.full_name
    db.session.delete(staff_member)
    db.session.commit()
    log_activity('Staff removed from campus', f'{full_name} removed from campus #{campus_id}', user=current_user)
    flash(f'{full_name} removed.', 'warning')
    return redirect(url_for('super_admin_campus_staff', campus_id=campus_id))


# ============ SEED DATA ============

DEFAULT_O_LEVEL_SUBJECTS = [
    ('Mathematics', 'MATH', 'O'), ('English Language', 'ENG', 'O'), ('Shona', 'SHO', 'O'),
    ('Ndebele', 'NDEB', 'O'), ('Combined Science', 'CSC', 'O'), ('Biology', 'BIO', 'O'),
    ('Chemistry', 'CHEM', 'O'), ('Physics', 'PHY', 'O'), ('History', 'HIST', 'O'),
    ('Geography', 'GEOG', 'O'), ('Accounts', 'ACCT', 'O'), ('Business Studies', 'BS', 'O'),
    ('Economics', 'ECON', 'O'), ('Computer Science', 'COMP', 'O'), ('Food & Nutrition', 'FOOD', 'O'),
    ('Fashion & Fabrics', 'FASH', 'O'), ('Woodwork', 'WOOD', 'O'), ('Metalwork', 'MET', 'O'),
    ('Technical Graphics', 'TECHG', 'O'), ('Building Studies', 'BUILD', 'O'), ('Agriculture', 'AGRI', 'O'),
    ('Religious Studies', 'RELS', 'O'), ('Literature in English', 'LITE', 'O'), ('French', 'FREN', 'O'),
    ('Portuguese', 'PORT', 'O'), ('Music', 'MUS', 'O'), ('Art', 'ART', 'O'),
]

DEFAULT_A_LEVEL_SUBJECTS = [
    ('Pure Mathematics', 'MATH-PURE', 'A'), ('Statistics', 'MATH-STAT', 'A'),
    ('Mechanics', 'MATH-MECH', 'A'), ('Further Mathematics', 'MATH-FURTHER', 'A'),
    ('English Literature', 'ELIT-A', 'A'), ('Shona', 'SHO-A', 'A'),
    ('Ndebele', 'NDEB-A', 'A'), ('Biology', 'BIO-A', 'A'), ('Chemistry', 'CHEM-A', 'A'),
    ('Physics', 'PHY-A', 'A'), ('History', 'HIST-A', 'A'), ('Geography', 'GEOG-A', 'A'),
    ('Accounts', 'ACCT-A', 'A'), ('Economics', 'ECON-A', 'A'), ('Business Studies', 'BS-A', 'A'),
    ('Computer Science', 'COMP-A', 'A'), ('Food Science', 'FOOD-A', 'A'), ('Fashion & Fabrics', 'FASH-A', 'A'),
    ('Technical Graphics', 'TECHG-A', 'A'), ('Agriculture', 'AGRI-A', 'A'), ('Divinity', 'DIV-A', 'A'),
    ('Religious Studies', 'RELS-A', 'A'), ('Literature in English', 'LITE-A', 'A'), ('French', 'FREN-A', 'A'),
    ('Portuguese', 'PORT-A', 'A'), ('Sociology', 'SOC-A', 'A'), ('Management of Business', 'MOB-A', 'A'),
]


def seed_campus_subjects(campus_id):
    """Create the standard O/A level subjects for a campus if it has none."""
    if Subject.query.filter_by(campus_id=campus_id).count() > 0:
        return True
    try:
        for name, code, level in DEFAULT_O_LEVEL_SUBJECTS + DEFAULT_A_LEVEL_SUBJECTS:
            db.session.add(Subject(name=name, code=code, level=level, campus_id=campus_id))
        db.session.commit()
        return True
    except IntegrityError:
        # Legacy global unique index/constraint still present (deploy lag).
        # Drop it on demand and retry so this page can never 500 on seed failure.
        db.session.rollback()
        try:
            if db.engine.dialect.name == 'postgresql':
                _migrate_table_uniques('subjects', 'code', 'uq_subjects_campus_code')
            for name, code, level in DEFAULT_O_LEVEL_SUBJECTS + DEFAULT_A_LEVEL_SUBJECTS:
                db.session.add(Subject(name=name, code=code, level=level, campus_id=campus_id))
            db.session.commit()
            return True
        except Exception as exc:
            db.session.rollback()
            print(f'[seed] retry failed after legacy drop: {exc}')
            return False

@app.route('/setup')
def setup():
    if Subject.query.first() is not None and os.environ.get('ENABLE_SETUP', '').lower() != 'true':
        return 'Setup is disabled.', 403
    db.drop_all()
    db.create_all()

    main_campus = Campus(name='Main Campus', code='MAIN', address='123 School Road, Harare')
    db.session.add(main_campus)
    db.session.commit()
    cid = main_campus.id

    seed_campus_subjects(cid)

    default_fees = [('Form 1', 350), ('Form 2', 350), ('Form 3', 400), ('Form 4', 500), ('Form 5', 450), ('Form 6', 600)]
    for form, fee in default_fees:
        db.session.add(FeeSetting(form=form, term_fee=fee, campus_id=cid))
    db.session.commit()

    principal = User(username='principal', email='principal@schoolbridge.zw', role='principal',
                     full_name='Dr. T. Chigumba', reg_number='REG-PRINCIPAL', phone='+263 712 345 678',
                     campus_id=cid)
    principal.set_password('principal123')
    db.session.add(principal)

    admin = User(username='admin', email='admin@schoolbridge.zw', role='admin',
                 full_name='Mr. K. Moyo', reg_number='REG-ADMIN', phone='+263 712 345 679',
                 campus_id=cid)
    admin.set_password('admin123')
    db.session.add(admin)

    cashier = User(username='cashier', email='cashier@schoolbridge.zw', role='cashier',
                   full_name='Mrs. S. Dube', reg_number='REG-CASHIER', phone='+263 712 345 680',
                   campus_id=cid)
    cashier.set_password('cashier123')
    db.session.add(cashier)

    super_admin = User(username='superadmin', email='superadmin@schoolbridge.zw', role='super_admin',
                       full_name='System Administrator', reg_number='REG-SUPERADMIN', phone='+263 712 345 687',
                       campus_id=cid)
    super_admin.set_password('superadmin123')
    db.session.add(super_admin)

    math_s = Subject.query.filter_by(code='MATH').first()
    eng_s = Subject.query.filter_by(code='ENG').first()
    sci_s = Subject.query.filter_by(code='CSC').first()
    bio_s = Subject.query.filter_by(code='BIO').first()
    hist_s = Subject.query.filter_by(code='HIST').first()
    geo_s = Subject.query.filter_by(code='GEOG').first()

    teachers_data = [
        ('teacher1', 'teacher1@schoolbridge.zw', 'Mr. A. Chikwanha', math_s.id, '+263 712 345 681'),
        ('teacher2', 'teacher2@schoolbridge.zw', 'Ms. P. Nyoni', eng_s.id, '+263 712 345 682'),
        ('teacher3', 'teacher3@schoolbridge.zw', 'Mr. B. Sibanda', sci_s.id, '+263 712 345 683'),
        ('teacher4', 'teacher4@schoolbridge.zw', 'Dr. M. Gumbo', bio_s.id, '+263 712 345 684'),
        ('teacher5', 'teacher5@schoolbridge.zw', 'Mrs. T. Moyo', hist_s.id, '+263 712 345 685'),
        ('teacher6', 'teacher6@schoolbridge.zw', 'Mr. C. Dlamini', geo_s.id, '+263 712 345 686'),
    ]
    for uname, email, name, subj_id, phone in teachers_data:
        t = User(username=uname, email=email, role='teacher', full_name=name,
                 subject_id=subj_id, phone=phone, reg_number=f'REG-{uname.upper()}',
                 campus_id=cid)
        t.set_password('teacher123')
        db.session.add(t)
    db.session.commit()

    forms = ['Form 4', 'Form 4', 'Form 3', 'Form 3', 'Form 4', 'Form 6']
    sample_data = [
        ('2026001', 'Tendai', 'Mukaro', 'Form 4', [math_s, eng_s, sci_s, bio_s, hist_s], '+263 712 000 001'),
        ('2026002', 'Chipo', 'Dube', 'Form 4', [math_s, eng_s, bio_s, geo_s], '+263 712 000 002'),
        ('2026003', 'Tafadzwa', 'Sithole', 'Form 3', [math_s, sci_s, geo_s, hist_s], '+263 712 000 003'),
        ('2026004', 'Rutendo', 'Gumbo', 'Form 3', [eng_s, sci_s, bio_s], '+263 712 000 004'),
        ('2026005', 'Kudzai', 'Zhou', 'Form 4', [math_s, eng_s, sci_s, bio_s, hist_s, geo_s], '+263 712 000 005'),
        ('2026006', 'Tanaka', 'Chigumba', 'Form 6', [Subject.query.filter_by(code='MATH-PURE').first(), Subject.query.filter_by(code='BIO-A').first(), Subject.query.filter_by(code='CHEM-A').first()], '+263 712 000 006'),
    ]

    for sid, fn, ln, form, subjs, phone in sample_data:
        reg = f'REG-{sid}'
        curriculum = 'Cambridge' if form in ('Form 5', 'Form 6') else 'ZIMSEC'
        student = Student(student_id=sid, first_name=fn, last_name=ln, form=form,
                         curriculum=curriculum,
                         email=f'{fn.lower()}.{ln.lower()}@student.schoolbridge.zw',
                         reg_number=reg, phone=phone, campus_id=cid)
        student.set_password('student123')
        db.session.add(student)
        db.session.flush()
        for subj in subjs:
            db.session.add(StudentSubject(student_id=student.id, subject_id=subj.id))

        balance = 0.0 if sid == '2026005' else 200.0
        fee = FeeAccount(student_id=student.id, campus_id=cid, term='Term 1', total_fees=500.00, amount_paid=500.00 - balance, balance=balance)
        db.session.add(fee)

    db.session.commit()

    # Sample marks for 2026001
    s1 = Student.query.filter_by(student_id='2026001').first()
    t1 = User.query.filter_by(username='teacher1').first()
    if s1 and t1:
        for month, mks in [('January', 68), ('February', 72), ('March', 65)]:
            db.session.add(MonthlyTest(student_id=s1.id, subject_id=math_s.id, term='Term 1',
                                       campus_id=cid,
                                       month=month, marks=mks, total_marks=100, teacher_id=t1.id))
        db.session.add(ExamMark(student_id=s1.id, subject_id=math_s.id, term='Term 1',
                                campus_id=cid,
                                exam_type='Mid-Term', marks=71, total_marks=100, teacher_id=t1.id))

    db.session.commit()
    return 'System setup complete! <a href="/auth/login">Login</a>'


def _migrate_a_level_math_subjects():
    campus = Campus.query.filter_by(code='MAIN').first()
    cid = campus.id if campus else None
    a_math = Subject.query.filter_by(code='MATH-A').first()
    if not a_math:
        a_math = Subject.query.filter_by(level='A', name='Mathematics').first()
    if a_math and not Subject.query.filter_by(code='MATH-PURE').first():
        a_math.name = 'Pure Mathematics'
        a_math.code = 'MATH-PURE'
    existing = {s.name for s in Subject.query.filter_by(level='A').all()}
    for name, code in (('Statistics', 'MATH-STAT'), ('Mechanics', 'MATH-MECH'), ('Further Mathematics', 'MATH-FURTHER')):
        if name not in existing:
            db.session.add(Subject(name=name, code=code, level='A', campus_id=cid))
    db.session.commit()


with app.app_context():
    try:
        _migrate_a_level_math_subjects()
    except Exception as exc:
        db.session.rollback()
        print(f'[migration] A-Level math subject update skipped: {exc}')


# ============ SECURITY ============

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "img-src 'self' data:; "
        "font-src 'self' data: https://cdn.jsdelivr.net; "
        "connect-src 'self'; "
        "frame-ancestors 'self'; "
        "form-action 'self'"
    )
    if request.is_secure:
        response.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    return response


COMMON_WEAK_PASSWORDS = {
    'password', 'password1', 'password123', '123456', '1234567', '12345678', '123456789',
    '1234567890', '123123', 'qwerty', 'qwerty123', 'abc123', 'admin', 'admin123',
    'administrator', 'letmein', 'welcome', 'iloveyou', 'monkey', 'dragon', 'master',
    'login', 'princess', 'football', 'shadow', 'sunshine', 'trustno1', 'default',
    'student123', 'teacher123', 'principal123', 'cashier123', 'school123', 'schoolbridge',
}


def password_strength_errors(password):
    """Return a list of problems with the password, or [] if it is strong."""
    errors = []
    if not password:
        errors.append('Password is required.')
        return errors
    if len(password) < 8:
        errors.append('Password must be at least 8 characters long.')
    if password.lower() in COMMON_WEAK_PASSWORDS:
        errors.append('That password is too common. Please choose a stronger one.')
    if not re.search(r'[A-Za-z]', password) or not re.search(r'\d', password):
        errors.append('Password must contain at least one letter and one number.')
    return errors


def is_valid_email(email):
    try:
        validate_email(email or '', check_deliverability=False)
        return True
    except EmailNotValidError:
        return False


def parse_float_field(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


@app.errorhandler(500)
def _debug_500(e):
    tb = traceback.format_exc()
    app.logger.error('500 error:\n%s', tb)
    return f'<html><body><h1>Internal Server Error</h1><pre>{tb}</pre></body></html>', 500


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=os.environ.get('FLASK_DEBUG', '').lower() == 'true', port=5000)
